from __future__ import print_function
import torch
import torch.nn as nn
from tqdm import tqdm
import optim
import util
import pdb
import numpy as np
from lifelines.utils import concordance_index
from evaluator.average_meter import AverageMeter
import pandas as pd
from openpyxl import load_workbook

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

class ModelEvaluator_test_xcal(object):
    """Class for evaluating a model during training"""
    def __init__(self, args, data_loaders, epochs_per_eval=1):
        """
        Args:
            data_loaders: List of Torch 'DataLoader's to sample from.
            num_visuals: Number of visuals to display from the validation set.
            max_eval: Maximum number of examples to evaluate at each evaluation.
            epochs_per_eval: Number of epochs between each evaluations.
        """
        self.args = args
        self.dataset = args.dataset
        self.data_loaders = data_loaders
        self.epochs_per_eval = epochs_per_eval
        self.loss_fn = optim.get_loss_fn(args.loss_fn, args)
        self.name = args.name
        self.lam = args.lam
        self.pred_type = args.pred_type
        self.model_dist = args.model_dist
        self.num_cat_bins = args.num_cat_bins
        self.loss_fn_name = args.loss_fn
        self.num_xcal_bins = args.num_xcal_bins
        if self.model_dist in ['cat', 'mtlr']:
            self.mid_points = args.mid_points
            self.bin_boundaries = args.bin_boundaries

    def evaluate(self, model, device, epoch=None):
        """
        Evaluate a model at the end of the given epoch.

        Args:
            model: Model to evaluate.
            device: Device on which to evaluate the model.
            epoch: The epoch that just finished. Determines whether to evaluate the model.

        Returns:
            metrics: Dictionary of metrics for the current model.

        Notes:
            Returned dictionary will be empty if not an evaluation epoch.
        """
        metrics = {}

        if epoch is None or epoch % self.epochs_per_eval == 0:
            # Evaluate on the training and validation sets
            model.eval()
            for data_loader in self.data_loaders:
                phase_metrics = self._eval_phase(model, data_loader, data_loader.phase, device)
                metrics.update(phase_metrics)
            model.train()

        return metrics

    def _eval_phase(self, model, data_loader, phase, device):
        print("CURRENT EVAL PHASE IS ", phase)
        """Evaluate a model for a single phase.

        Args:
            model: Model to evaluate.
            data_loader: Torch DataLoader to sample from.
            phase: Phase being evaluated. One of 'train', 'val', or 'test'.
            device: Device on which to evaluate the model.

        Returns:
            metrics: Dictionary of metrics for the phase.
        """
        # Keep track of task-specific records needed for computing overall metrics
        records = {'loss_meter': AverageMeter()}
        
        num_examples = len(data_loader.dataset)
        # Sample from the data loader and record model outputs
        loss_fn = self.loss_fn
        num_evaluated = 0

        is_dead_per_batch = []

        all_cdf = []
        all_tte = []
        for src, tgt in data_loader:
            src = src.to(DEVICE)
            tgt = tgt.to(DEVICE)

            tte = tgt[:, 0]

            if self.model_dist in ['cat', 'mtlr']:
                tgt = util.cat_bin_target(self.args, tgt, self.bin_boundaries) # check again
                
            # THIS MUST COME AFTER THE ABOCE CAT BIN TARGET FUNCTION
            # BECAUSE CAT BIN TARGET CAN CHANGE SOME IS_DEAD
            is_dead = tgt[:, 1]
            order = torch.argsort(tte)
            is_dead = is_dead[order]

            is_dead_per_batch.append(is_dead)

            if num_evaluated >= num_examples:
                break

            pred_params = model.forward(src)
            cdf = util.get_cdf_val(pred_params, tgt, self.args) # check again
            all_cdf.append(cdf)
            loss = loss_fn(pred_params, tgt, model_dist=self.model_dist)
            num_both = pred_params.size()[0]
            all_tte.append(tte)

            self._record_batch(num_both, loss, **records)
            num_evaluated += src.size(0)
            
        #if self.args.dataset in ['nacd', 'nacdcol', 'brca', 'read', 'gbm', 'gbmlgg', 'dbcd', 'dlbcl']:
        #    concordance = util.concordance(self.args, data_loader, model)

        #else:
        #    concordance = -1.0
        concordance = util.concordance(self.args, data_loader, model)
        
        is_dead = torch.cat(is_dead_per_batch).long()
        
        all_cdf = torch.cat(all_cdf)

        all_tte = torch.cat(all_tte)
        #torch.save(all_tte, "C:/Users/wjdgh/Desktop/collection/xcal" + '_' + str(self.args.k) + '_' + str(self.args.lam) + "_tte.pt")
        #torch.save(is_dead, "C:/Users/wjdgh/Desktop/collection/xcal" + '_' + str(self.args.k) + '_' + str(self.args.lam) + "_is_dead.pt")
        #torch.save(all_cdf, "C:/Users/wjdgh/Desktop/collection/xcal" + '_' + str(self.args.k) + '_' + str(self.args.lam) + "_cdf.pt")
        if self.args.model_dist == 'mtlr':
            weight = model.get_weight()
            regularizer = util.ridge_norm(weight)*self.args.C1/2 + util.fused_norm(weight)*self.args.C2/2

        # Map to summary dictionaries
        metrics = self._get_summary_dict(phase, **records)
        approx_s_calibration = util.s_calibration(points=all_cdf, is_dead=is_dead, args=self.args, gamma=1e5, differentiable=False, device=DEVICE)
        test_statistic, p_value = util.get_p_value(cdf=all_cdf, tte=all_tte, is_dead=is_dead, device=DEVICE) # check again
        approx_d_calibration_10 = util.d_calibration(points=all_cdf, is_dead=is_dead, args=self.args, nbins=10, gamma=1e5, differentiable=False, device=DEVICE)
        approx_d_calibration_20 = util.d_calibration(points=all_cdf, is_dead=is_dead, args=self.args, nbins=self.num_xcal_bins, gamma=1e5, differentiable=False, device=DEVICE)
        approx_d_calibration_40 = util.d_calibration(points=all_cdf, is_dead=is_dead, args=self.args, nbins=self.num_xcal_bins*2, gamma=1e5, differentiable=False, device=DEVICE)
        approx_d_calibration_60 = util.d_calibration(points=all_cdf, is_dead=is_dead, args=self.args, nbins=self.num_xcal_bins*3, gamma=1e5, differentiable=False, device=DEVICE)
        
        metrics[phase + '_' + 'NLL'] = metrics[phase + '_' + 'loss']
        metrics[phase + '_' + 'concordance'] = concordance
        metrics[phase + '_' + 'scal(20)'] = approx_s_calibration
        metrics[phase + '_' + 'dcal(10)'] = approx_d_calibration_10
        metrics[phase + '_' + 'dcal(20)'] = approx_d_calibration_20
        metrics[phase + '_' + 'dcal(40)'] = approx_d_calibration_40
        metrics[phase + '_' + 'dcal(60)'] = approx_d_calibration_60
        if self.model_dist in ['mtlr']:
            metrics[phase + '_' + 'loss'] = metrics[phase + '_' + 'loss'] + self.lam * approx_d_calibration_20 + regularizer

        else:
            metrics[phase + '_' + 'loss'] = metrics[phase + '_' + 'loss'] + self.lam * approx_d_calibration_20
        metrics[phase + '_' + 'teststat'] = test_statistic
        metrics[phase + '_' + 'pvalue'] = p_value
        #metrics[phase + '_' + 'brier score'] = brier_score
        
        if phase == 'test':
            workbook = load_workbook(filename='C:/Users/wjdgh/Desktop/tmp.xlsx')
            sheet = workbook.active
            last_row = sheet.max_row
            sheet.cell(row=last_row+1, column=1, value=metrics[phase + '_' + 'loss'].item())
            sheet.cell(row=last_row+1, column=2, value=metrics[phase + '_' + 'NLL'])
            sheet.cell(row=last_row+1, column=3, value=metrics[phase + '_' + 'concordance'].item())
            sheet.cell(row=last_row+1, column=4, value=metrics[phase + '_' + 'scal(20)'].item())
            sheet.cell(row=last_row+1, column=5, value=metrics[phase + '_' + 'dcal(10)'].item())
            sheet.cell(row=last_row+1, column=6, value=metrics[phase + '_' + 'dcal(20)'].item())
            sheet.cell(row=last_row+1, column=7, value=metrics[phase + '_' + 'dcal(40)'].item())
            sheet.cell(row=last_row+1, column=8, value=metrics[phase + '_' + 'dcal(60)'].item())
            sheet.cell(row=last_row+1, column=9, value=metrics[phase + '_' + 'teststat'].item())
            sheet.cell(row=last_row+1, column=10, value=metrics[phase + '_' + 'pvalue'].item())
            workbook.save('C:/Users/wjdgh/Desktop/tmp.xlsx')
        
        print(' ---- {} epoch Concordance {:.4f}'.format(phase, concordance))
        print(' ---- {} epoch end S-cal(20) {:.5f}'.format(phase, approx_s_calibration))
        print(' ---- {} epoch end D-cal(10) {:.5f}'.format(phase, approx_d_calibration_10))
        print(' ---- {} epoch end D-cal(20) {:.5f}'.format(phase, approx_d_calibration_20))
        print(' ---- {} epoch end D-cal(40) {:.5f}'.format(phase, approx_d_calibration_40))
        print(' ---- {} epoch end D-cal(60) {:.5f}'.format(phase, approx_d_calibration_60))

        print(torch.histogram(all_cdf.to('cpu'), bins=20, range=(0, 1)))
        
        return metrics

    @staticmethod
    def _record_batch(N, loss, loss_meter=None):
        """
        Record results from a batch to keep track of metrics during evaluation.

        Args:
            logits: Batch of logits output by the model.
            loss_meter: AverageMeter keeping track of average loss during evaluation.
        """
        if loss_meter is not None:
            loss_meter.update(loss.item(), N)

    @staticmethod
    def _get_summary_dict(phase, loss_meter=None):
        """
        Get summary dictionaries given dictionary of records kept during evaluation.

        Args:
            phase: Phase being evaluated. One of 'train', 'val', or 'test'.
            loss_meter: AverageMeter keeping track of average loss during evaluation.

        Returns:
            metrics: Dictionary of metrics for the current model.
        """
        metrics = {phase + '_' + 'loss': loss_meter.avg}
        
        return metrics

    @staticmethod
    def _write_summary_stats(phase, loss_meter=None):
        """
        Write stats of evaluation to file.

        Args:
            phase: Phase being evaluated. One of 'train', 'val', or 'test'.
            loss_meter: AverageMeter keeping track of average loss during evaluation.

        Returns:
            metrics: Dictionary of metrics for the current model.
        """
        metrics = {phase + '_' + 'loss': loss_meter.avg}

        return metrics
